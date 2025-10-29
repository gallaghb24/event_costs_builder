import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import os
import subprocess
import json
from datetime import datetime
import io
import tempfile
from typing import Dict, List, Tuple, Optional
import re
import math

# Page config
st.set_page_config(
    page_title="Superdrug ITG Invoice Generator",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'production_data' not in st.session_state:
    st.session_state.production_data = pd.DataFrame()
if 'studio_data' not in st.session_state:
    st.session_state.studio_data = pd.DataFrame()
if 'print_data' not in st.session_state:
    st.session_state.print_data = pd.DataFrame()
if 'timesheet_hours' not in st.session_state:
    st.session_state.timesheet_hours = pd.DataFrame()
if 'template_loaded' not in st.session_state:
    st.session_state.template_loaded = False
if 'generated_file' not in st.session_state:
    st.session_state.generated_file = None

def round_up_to_quarter(hours):
    """Round hours up to nearest 0.25"""
    if pd.isna(hours) or hours == 0:
        return 0.0
    return math.ceil(hours * 4) / 4

def convert_event_to_code(event_name: str) -> str:
    """Convert 'Event 10 2025' to 'E1025' format"""
    match = re.search(r'Event\s+(\d+)\s+(\d{4})', event_name, re.IGNORECASE)
    if match:
        event_num = match.group(1).zfill(2)
        year = match.group(2)[-2:]
        return f"E{event_num}{year}"
    return "E0000"

def process_timesheet(file) -> pd.DataFrame:
    """Process timesheet CSV to extract and aggregate studio hours"""
    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_file:
        tmp_file.write(file.getbuffer())
        tmp_path = tmp_file.name
    
    # Read the timesheet with encoding fallbacks
    encodings_to_try = ["utf-8-sig", "utf-16", "latin1", "cp1252"]
    timesheet_df = None
    last_exception = None

    try:
        for encoding in encodings_to_try:
            try:
                timesheet_df = pd.read_csv(tmp_path, encoding=encoding)
                break
            except UnicodeDecodeError as exc:
                last_exception = exc
            except Exception as exc:
                last_exception = exc
                break
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    if timesheet_df is None:
        if isinstance(last_exception, UnicodeDecodeError):
            st.error(
                "Unable to read the uploaded timesheet. Please upload a CSV encoded as UTF-8 or UTF-16."
            )
        elif last_exception is not None:
            st.error(f"Failed to read the uploaded timesheet: {last_exception}")
        else:
            st.error("Unable to read the uploaded timesheet.")
        return pd.DataFrame()
    
    # Extract job number prefix (e.g., SDG2161 from 1/SDG2161)
    timesheet_df['Job_Prefix'] = timesheet_df['Job Number'].str.extract(r'1/(SDG\d+)')
    
    # Filter out Studio QC hours
    if 'Charge Code' in timesheet_df.columns:
        # Exclude rows where Charge Code contains 'QC' or 'Studio QC'
        mask = ~timesheet_df['Charge Code'].str.contains('QC', case=False, na=False)
        timesheet_df = timesheet_df[mask]
    
    # Filter for SDG jobs
    sdg_jobs = timesheet_df[timesheet_df['Job_Prefix'].notna()]
    
    if sdg_jobs.empty:
        return pd.DataFrame()
    
    # Aggregate hours by job
    aggregated = sdg_jobs.groupby('Job_Prefix').agg({
        'Job Description': 'first',
        'Total': 'sum',  # Sum of all hours (excluding QC)
        'Charge Code': lambda x: x.mode()[0] if not x.empty else '',
    }).reset_index()
    
    # Rename columns
    aggregated.columns = ['Project Ref', 'Project Description', 'Total Hours', 'Primary Charge Code']
    
    # Round hours up to nearest 0.25
    aggregated['Total Hours'] = aggregated['Total Hours'].apply(round_up_to_quarter)
    
    # Determine Type based on Charge Code
    def determine_type(charge_code):
        if pd.isna(charge_code):
            return 'Artwork'
        charge_code_lower = str(charge_code).lower()
        if 'creative' in charge_code_lower:
            return 'Creative Artwork'
        elif 'digital' in charge_code_lower or 'tec' in charge_code_lower:
            return 'Digital'
        else:
            return 'Artwork'
    
    aggregated['Type'] = aggregated['Primary Charge Code'].apply(determine_type)
    
    # Set default Core/OAB based on description
    aggregated['Core/OAB'] = 'CORE'
    roi_mask = aggregated['Project Description'].str.contains('ROI', case=False, na=False)
    aggregated.loc[roi_mask, 'Core/OAB'] = 'OAB'
    
    return aggregated[['Project Ref', 'Total Hours', 'Type', 'Core/OAB']]

def load_template(file) -> Dict:
    """Load and analyze the Excel template"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(file.getbuffer())
        tmp_path = tmp_file.name
    
    wb = load_workbook(tmp_path, data_only=False, keep_vba=True)

    # Store formatting information for each sheet
    formatting_info = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_format = {
            'column_widths': {},
            'row_heights': {},
            'merged_cells': list(sheet.merged_cells.ranges),
            'cell_styles': {}
        }
        
        # Store column widths
        for col in sheet.column_dimensions:
            sheet_format['column_widths'][col] = sheet.column_dimensions[col].width
        
        # Store row heights
        for row in sheet.row_dimensions:
            sheet_format['row_heights'][row] = sheet.row_dimensions[row].height
        
        # Store cell styles for header rows (first 3 rows)
        for row in range(1, 4):
            for col in range(1, min(sheet.max_column + 1, 50)):
                cell = sheet.cell(row=row, column=col)
                if cell.value or cell.font or cell.fill or cell.border:
                    coord = cell.coordinate
                    sheet_format['cell_styles'][coord] = {
                        'font': cell.font.copy() if cell.font else None,
                        'fill': cell.fill.copy() if cell.fill else None,
                        'border': cell.border.copy() if cell.border else None,
                        'alignment': cell.alignment.copy() if cell.alignment else None,
                        'number_format': cell.number_format
                    }
        
        formatting_info[sheet_name] = sheet_format
    
    # Extract key information
    template_info = {
        'path': tmp_path,
        'sheets': wb.sheetnames,
        'core_clients': [],
        'oab_clients': [],
        'studio_types': ['Artwork', 'Creative Artwork', 'Digital'],
        'wb': wb,
        'formatting': formatting_info,
        'has_macros': bool(getattr(wb, 'vba_archive', None)),
    }
    
    # Extract client names from Event Summary sheets
    for sheet_name in ['Event Summary - Core', 'Event Summary - OAB']:
        sheet = wb[sheet_name]
        clients = []
        for row in range(7, 50):
            cell_b = sheet[f'B{row}'].value
            if cell_b and cell_b not in ['Total', 'TOTAL', ''] and not str(cell_b).startswith('='):
                clients.append(cell_b)
        
        if 'Core' in sheet_name:
            template_info['core_clients'] = clients
        else:
            template_info['oab_clients'] = clients
    
    return template_info

def load_production_files(files) -> pd.DataFrame:
    """Load and combine production files, handling deduplication"""
    all_data = []
    
    for file in files:
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(file.getbuffer())
            tmp_path = tmp_file.name
        
        # Read with header at row 2 (index 1)
        df = pd.read_excel(tmp_path, header=1)
        all_data.append(df)
        os.remove(tmp_path)
    
    # Combine all dataframes
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # Remove duplicates based on Brief Ref
    combined_df = combined_df.drop_duplicates(subset=['Brief Ref'], keep='first')
    
    return combined_df

def filter_production_data(df: pd.DataFrame) -> pd.DataFrame:
    """Apply production data enhancements without removing rows."""
    if df.empty:
        return df

    # Define statuses that previously triggered exclusion
    exclude_statuses = {
        'draft', 'saved', 'awaiting rfq', 'rfq responses',
        'estimates awaiting approval', 'client approved estimates'
    }

    df = df.copy()

    if 'Production Supplier Brief Status' in df.columns:
        status_series = df['Production Supplier Brief Status']
        status_stripped = status_series.apply(lambda x: x.strip() if isinstance(x, str) else x)
        status_lower = status_stripped.fillna('').str.lower()

        df['Production Supplier Brief Status'] = status_stripped
        df['Production Status Note'] = np.where(
            status_lower.isin(exclude_statuses),
            'check status/cost as line not in production yet',
            ''
        )
    else:
        df['Production Status Note'] = ''

    return df

def prepare_print_data(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare data for the Print tab"""
    # Return the dataframe with all original columns for correct mapping
    return df.copy()

def prepare_studio_data(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare data for the Studio tab - aggregate at job level"""
    base_columns = [
        'Project Ref', 'Event Name', 'Project Description', 'Project Owner',
        'Lines', 'Studio Hours', 'Type', 'Core/OAB', 'Studio Comment'
    ]

    if df.empty or 'Project Ref' not in df.columns:
        return pd.DataFrame(columns=base_columns)

    comment_text = 'check all lines are approved, artwork hours may require updating'

    working_df = df.copy()

    if 'Content Brief Status' in working_df.columns:
        status_series = working_df['Content Brief Status']
        status_stripped = status_series.apply(lambda x: x.strip() if isinstance(x, str) else x)
        status_lower = status_stripped.fillna('').str.lower()
        working_df['Content Brief Status'] = status_stripped
    else:
        status_lower = pd.Series([''] * len(working_df), index=working_df.index, dtype='object')

    working_df['__status_lower'] = status_lower

    status_groups = working_df.groupby('Project Ref')['__status_lower'].apply(list)

    keep_projects = []
    comment_projects = set()

    for project_ref, statuses in status_groups.items():
        if not statuses:
            keep_projects.append(project_ref)
            comment_projects.add(project_ref)
            continue

        unique_statuses = set(statuses)
        if unique_statuses and unique_statuses <= {'not applicable'}:
            continue  # Skip projects that are entirely "not applicable"

        keep_projects.append(project_ref)

        if any(status != 'completed' for status in statuses):
            comment_projects.add(project_ref)

    if not keep_projects:
        return pd.DataFrame(columns=base_columns)

    filtered_df = working_df[working_df['Project Ref'].isin(keep_projects)].copy()

    valid_rows = filtered_df[filtered_df['__status_lower'] != 'not applicable'].copy()

    if valid_rows.empty:
        return pd.DataFrame(columns=base_columns)

    grouped = valid_rows.groupby('Project Ref').agg({
        'Event Name': 'first',
        'Project Description': 'first',
        'Project Owner': 'first',
        'Brief Ref': 'count'  # Count of lines per job
    }).reset_index()

    grouped.columns = ['Project Ref', 'Event Name', 'Project Description', 'Project Owner', 'Lines']

    grouped['Studio Hours'] = None  # Will be filled from timesheet
    grouped['Type'] = ''  # Will be filled from timesheet
    grouped['Core/OAB'] = ''  # Will be filled from timesheet
    grouped['Studio Comment'] = ''

    if comment_projects:
        grouped.loc[grouped['Project Ref'].isin(comment_projects), 'Studio Comment'] = comment_text

    return grouped

def apply_formatting(sheet, formatting_info):
    """Apply saved formatting to a worksheet"""
    if not formatting_info:
        return
    
    # Apply column widths
    for col, width in formatting_info.get('column_widths', {}).items():
        sheet.column_dimensions[col].width = width
    
    # Apply row heights
    for row, height in formatting_info.get('row_heights', {}).items():
        sheet.row_dimensions[row].height = height
    
    # Apply cell styles
    for coord, style in formatting_info.get('cell_styles', {}).items():
        try:
            cell = sheet[coord]
            if style['font']:
                cell.font = style['font']
            if style['fill']:
                cell.fill = style['fill']
            if style['border']:
                cell.border = style['border']
            if style['alignment']:
                cell.alignment = style['alignment']
            if style['number_format']:
                cell.number_format = style['number_format']
        except:
            pass  # Skip if cell doesn't exist
    
    # Restore merged cells
    for merged_range in formatting_info.get('merged_cells', []):
        try:
            sheet.merge_cells(str(merged_range))
        except:
            pass  # Skip if merge fails

def generate_invoice(template_info: Dict, studio_df: pd.DataFrame, print_df: pd.DataFrame,
                     event_name: str, event_code: str) -> Tuple[str, str, str]:
    """Generate the invoice Excel file with proper formatting.

    Returns a tuple of (output_path, download_filename, mime_type).
    """
    
    # Load the template
    wb = load_workbook(template_info['path'], data_only=False, keep_vba=True)
    
    # Update Event Summary sheets with event name
    for sheet_name in ['Event Summary - Core', 'Event Summary - OAB']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet['D4'] = event_name
            
            # Preserve formatting
            if sheet_name in template_info['formatting']:
                apply_formatting(sheet, template_info['formatting'][sheet_name])
    
    # Populate Studio sheet
    if 'Studio' in wb.sheetnames and not studio_df.empty:
        sheet = wb['Studio']
        
        # Preserve formatting first
        if 'Studio' in template_info['formatting']:
            apply_formatting(sheet, template_info['formatting']['Studio'])
        
        # Clear existing data (keep headers in row 2)
        for row in range(3, min(sheet.max_row + 1, 1000)):
            for col in range(1, 15):
                cell = sheet.cell(row=row, column=col)
                cell.value = None
        
        # Add studio data
        for idx, (_, job) in enumerate(studio_df.iterrows(), start=3):
            sheet[f'A{idx}'] = job['Project Ref']
            sheet[f'B{idx}'] = job['Event Name']
            sheet[f'C{idx}'] = job['Project Description']
            sheet[f'D{idx}'] = job['Project Owner']
            sheet[f'E{idx}'] = job['Lines']

            # Add hours and type if available
            if pd.notna(job.get('Studio Hours')):
                sheet[f'F{idx}'] = job['Studio Hours']
            if job.get('Type'):
                sheet[f'G{idx}'] = job['Type']

            # Formula for rate
            sheet[f'H{idx}'] = f'=IF(G{idx}="Artwork",49.5,IF(G{idx}="Creative Artwork",57,IF(G{idx}="Digital",49.5,0)))'

            # Formula for cost
            sheet[f'I{idx}'] = f'=F{idx}*H{idx}'

            # Core/OAB if available
            if job.get('Core/OAB'):
                sheet[f'J{idx}'] = job['Core/OAB']

            # Add studio comment if applicable
            comment_value = job.get('Studio Comment', '')
            if pd.notna(comment_value):
                comment_text = str(comment_value).strip()
                if comment_text:
                    sheet[f'A{idx}'].comment = Comment(comment_text, "Status")
    
    # Populate Print sheet
    if 'Print' in wb.sheetnames and not print_df.empty:
        sheet = wb['Print']
        
        # Preserve formatting first
        if 'Print' in template_info['formatting']:
            apply_formatting(sheet, template_info['formatting']['Print'])
        
        # Clear existing data (keep headers in row 2)
        for row in range(3, min(sheet.max_row + 1, 3000)):
            for col in range(1, 30):  # Up to column AC
                cell = sheet.cell(row=row, column=col)
                cell.value = None
        
        # Add print data with CORRECT field mappings
        for idx, (_, item) in enumerate(print_df.iterrows(), start=3):
            # Columns A-K: Basic information
            sheet[f'A{idx}'] = item.get('Project Ref', '')
            sheet[f'B{idx}'] = item.get('Event Name', '')
            sheet[f'C{idx}'] = item.get('Project Description', '')
            sheet[f'D{idx}'] = item.get('Project Owner', '')
            sheet[f'E{idx}'] = item.get('Brief Ref', '')
            sheet[f'F{idx}'] = item.get('POS Code', '')
            sheet[f'G{idx}'] = item.get('Brief Description', '')
            sheet[f'H{idx}'] = str(item.get('Part URN', ''))
            sheet[f'I{idx}'] = item.get('Part', '')
            sheet[f'J{idx}'] = item.get('Height', 0)
            sheet[f'K{idx}'] = item.get('Width', 0)
            
            # Columns L-S: Production details
            sheet[f'L{idx}'] = item.get('Colours Front', '')
            sheet[f'M{idx}'] = item.get('Colours Back', '')
            sheet[f'N{idx}'] = item.get('Material', '')
            sheet[f'O{idx}'] = item.get('No of Pages', 0)  # No of Pages from input
            sheet[f'P{idx}'] = item.get('Production Finishing Notes', '')
            sheet[f'Q{idx}'] = item.get('Production Supplier Comments', '')
            sheet[f'R{idx}'] = item.get('Allocated Qty', 0)
            sheet[f'S{idx}'] = item.get('Spares', 0)
            
            # Columns T-Y: Status and costs
            sheet[f'T{idx}'] = item.get('Total including Spares', 0)  # Total including Spares
            sheet[f'U{idx}'] = item.get('No of Stores', 0)  # No of Stores
            sheet[f'V{idx}'] = str(item.get('In Store Deadline', ''))  # In Store Deadline
            sheet[f'W{idx}'] = item.get('Content Brief Status', '')  # Content Brief Status
            sheet[f'X{idx}'] = item.get('Production Supplier Brief Status', '')  # Production Supplier Brief Status
            sheet[f'Y{idx}'] = item.get('Production Sell Price', 0)  # Production Cost (from Sell Price)

            # Add production status comment when flagged
            status_note = item.get('Production Status Note', '')
            if pd.notna(status_note):
                status_note_text = str(status_note).strip()
                if status_note_text:
                    sheet[f'X{idx}'].comment = Comment(status_note_text, "Status")
            
            # Column Z: Core/OAB lookup formula
            sheet[f'Z{idx}'] = f'=IF(Y{idx}>0,IFERROR(VLOOKUP(A{idx},Studio!$A$3:$J$6129,10,FALSE),""),"")' 
            
            # Column AA: Comments if exists
            if 'Comments' in item:
                sheet[f'AA{idx}'] = item.get('Comments', '')
    
    # Preserve formatting for all other sheets
    for sheet_name in wb.sheetnames:
        if sheet_name not in ['Studio', 'Print', 'Event Summary - Core', 'Event Summary - OAB']:
            if sheet_name in template_info['formatting']:
                sheet = wb[sheet_name]
                apply_formatting(sheet, template_info['formatting'][sheet_name])
    
    # Save the file with event code in filename
    now = datetime.now()
    extension = '.xlsm' if template_info.get('has_macros') else '.xlsx'
    output_path = f"/tmp/{event_code}_Superdrug_ITG_Invoice_{now.strftime('%Y%m%d_%H%M%S')}{extension}"
    wb.save(output_path)
    wb.close()

    download_filename = f"{event_code}_Superdrug_ITG_Invoice_{now.strftime('%Y%m%d')}{extension}"
    mime_type = (
        'application/vnd.ms-excel.sheet.macroEnabled.12'
        if extension == '.xlsm'
        else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    return output_path, download_filename, mime_type

# Main UI
st.title("📊 Superdrug ITG Invoice Generator v3.0")
st.markdown("---")

# Sidebar for configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    
    # Template upload
    st.subheader("1. Upload Template")
    template_file = st.file_uploader(
        "Select Excel Template",
        type=['xlsx', 'xlsm'],
        help="Upload the Superdrug ITG Invoice Template"
    )
    
    if template_file:
        if st.button("Load Template"):
            with st.spinner("Loading template..."):
                st.session_state.template_info = load_template(template_file)
                st.session_state.template_loaded = True
                st.success("Template loaded successfully!")
    
    if st.session_state.template_loaded:
        st.success("✅ Template Loaded")
        with st.expander("Template Info"):
            info = st.session_state.template_info
            st.write(f"**Sheets:** {len(info['sheets'])}")
            st.write(f"**Core Clients:** {len(info['core_clients'])}")
            st.write(f"**OAB Clients:** {len(info['oab_clients'])}")
            st.write(f"**Has Macros:** {'Yes' if info.get('has_macros') else 'No'}")

# Main content area
if st.session_state.template_loaded:
    
    # Event Information
    st.header("📅 Event Information")
    col1, col2 = st.columns([2, 1])
    with col1:
        event_name = st.text_input(
            "Event", 
            value="Event 10 2025", 
            help="Enter the event name (e.g., Event 10 2025). This will be converted to a code like E1025 for the filename."
        )
    with col2:
        event_code = convert_event_to_code(event_name)
        st.text_input("Event Code (auto-generated)", value=event_code, disabled=True)
    
    # Tabs for different data inputs
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📁 Production Files", 
        "⏰ Timesheet Data",
        "👨‍💻 Studio Hours Review", 
        "🔍 Data Review", 
        "💰 Cost Preview", 
        "📄 Generate Invoice"
    ])
    
    with tab1:
        st.header("📁 Upload Production Files")
        st.markdown("""
        Upload the Production Lines INTERNAL Excel files. The app will:
        - Combine multiple files
        - Remove duplicates based on Brief Ref
        - Filter out draft/unapproved items
        - Prepare data for Studio and Print tabs
        """)
        
        production_files = st.file_uploader(
            "Select Production Files",
            type=['xlsx'],
            accept_multiple_files=True,
            help="Upload one or more Production Lines INTERNAL files"
        )
        
        if production_files:
            if st.button("Process Production Files", type="primary"):
                with st.spinner("Processing production files..."):
                    # Load and combine files
                    combined_df = load_production_files(production_files)
                    st.info(f"Loaded {len(combined_df)} total rows from {len(production_files)} file(s)")
                    
                    # Filter data
                    filtered_df = filter_production_data(combined_df)
                    st.info(f"After filtering: {len(filtered_df)} rows remaining")
                    
                    # Prepare Print data
                    print_df = prepare_print_data(filtered_df)
                    st.session_state.print_data = print_df
                    
                    # Prepare Studio data
                    studio_df = prepare_studio_data(filtered_df)
                    st.session_state.studio_data = studio_df
                    
                    # Store full data for reference
                    st.session_state.production_data = filtered_df
                    
                    st.success(f"✅ Processed successfully!")
                    st.write(f"- **Studio Jobs:** {len(studio_df)} unique projects")
                    st.write(f"- **Print Items:** {len(print_df)} line items")
        
        if not st.session_state.production_data.empty:
            st.subheader("📊 Processed Data Summary")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Projects", len(st.session_state.studio_data))
            with col2:
                st.metric("Total Print Lines", len(st.session_state.print_data))
            with col3:
                total_lines = st.session_state.studio_data['Lines'].sum()
                st.metric("Total Brief Lines", int(total_lines))
    
    with tab2:
        st.header("⏰ Upload Timesheet Data")
        
        if st.session_state.studio_data.empty:
            st.warning("⚠️ Please process Production Files first")
        else:
            st.markdown("""
            Upload the timesheet CSV to automatically populate Studio Hours. The app will:
            - **Exclude Studio QC hours** (not chargeable)
            - **Round hours up to nearest 0.25**
            - Match projects by Project Ref
            - Determine Type from charge codes
            - Set Core/OAB based on ROI designation
            """)
            
            timesheet_file = st.file_uploader(
                "Select Timesheet CSV",
                type=['csv'],
                help="Upload the timesheet export CSV file"
            )
            
            if timesheet_file:
                if st.button("Process Timesheet", type="primary"):
                    with st.spinner("Processing timesheet data..."):
                        # Process timesheet
                        timesheet_hours = process_timesheet(timesheet_file)
                        
                        if timesheet_hours.empty:
                            st.error("No valid timesheet data found")
                        else:
                            st.session_state.timesheet_hours = timesheet_hours
                            
                            # Merge with studio data
                            studio_df = st.session_state.studio_data.copy()
                            
                            # Merge timesheet hours
                            merged = studio_df.merge(
                                timesheet_hours,
                                on='Project Ref',
                                how='left',
                                suffixes=('', '_timesheet')
                            )
                            
                            # Update with timesheet data where available
                            merged['Studio Hours'] = merged['Total Hours'].fillna(merged['Studio Hours'])
                            merged['Type'] = merged['Type_timesheet'].fillna(merged['Type']).replace('', 'Artwork')
                            merged['Core/OAB'] = merged['Core/OAB_timesheet'].fillna(merged['Core/OAB']).replace('', 'CORE')
                            
                            # Drop timesheet columns
                            merged = merged.drop(['Total Hours', 'Type_timesheet', 'Core/OAB_timesheet'], axis=1, errors='ignore')
                            
                            # Update session state
                            st.session_state.studio_data = merged
                            
                            # Show summary
                            matched = merged['Studio Hours'].notna().sum()
                            total_hours = merged['Studio Hours'].sum()
                            
                            st.success(f"✅ Timesheet processed successfully!")
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Projects Matched", f"{matched}/{len(merged)}")
                            with col2:
                                st.metric("Total Hours", f"{total_hours:.2f}")
                            with col3:
                                avg_hours = total_hours / matched if matched > 0 else 0
                                st.metric("Avg Hours/Project", f"{avg_hours:.2f}")
                            
                            # Show unmatched projects
                            unmatched = merged[merged['Studio Hours'].isna()]
                            if len(unmatched) > 0:
                                with st.expander(f"⚠️ {len(unmatched)} projects without hours"):
                                    st.dataframe(unmatched[['Project Ref', 'Project Description', 'Lines']])
    
    with tab3:
        st.header("👨‍💻 Studio Hours Review & Edit")
        
        if not st.session_state.studio_data.empty:
            # Check if hours have been added
            has_hours = st.session_state.studio_data['Studio Hours'].notna().any()
            
            if not has_hours:
                st.info("💡 Upload a timesheet in the previous tab to auto-populate hours, or enter them manually below")
            
            st.markdown("Review and edit studio hours, types, and Core/OAB assignments as needed.")
            
            # Editable dataframe
            edit_df = st.session_state.studio_data.copy()
            edit_df['Studio Hours'] = edit_df['Studio Hours'].fillna(0.0)
            edit_df['Type'] = edit_df['Type'].replace('', 'Artwork')
            edit_df['Core/OAB'] = edit_df['Core/OAB'].replace('', 'CORE')
            if 'Studio Comment' not in edit_df.columns:
                edit_df['Studio Comment'] = ''
            edit_df['Studio Comment'] = edit_df['Studio Comment'].fillna('')

            # Create editable columns configuration
            column_config = {
                "Project Ref": st.column_config.TextColumn("Project Ref", disabled=True),
                "Event Name": st.column_config.TextColumn("Event Name", disabled=True),
                "Project Description": st.column_config.TextColumn("Description", disabled=True),
                "Project Owner": st.column_config.TextColumn("Owner", disabled=True),
                "Lines": st.column_config.NumberColumn("Lines", disabled=True, format="%d"),
                "Studio Hours": st.column_config.NumberColumn("Hours", min_value=0, max_value=1000, step=0.25, required=True),
                "Type": st.column_config.SelectboxColumn(
                    "Type",
                    options=["Artwork", "Creative Artwork", "Digital"],
                    default="Artwork",
                    required=True
                ),
                "Core/OAB": st.column_config.SelectboxColumn(
                    "Core/OAB",
                    options=["CORE", "OAB"],
                    default="CORE",
                    required=True
                )
            }

            if 'Studio Comment' in edit_df.columns:
                column_config["Studio Comment"] = st.column_config.TextColumn(
                    "Note",
                    disabled=True
                )
            
            edited_studio = st.data_editor(
                edit_df,
                column_config=column_config,
                use_container_width=True,
                num_rows="fixed",
                height=400
            )
            
            # Update session state with edited data
            st.session_state.studio_data = edited_studio
            
            # Show summary
            if edited_studio['Studio Hours'].sum() > 0:
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    total_hours = edited_studio['Studio Hours'].sum()
                    st.metric("Total Hours", f"{total_hours:.2f}")
                with col2:
                    core_hours = edited_studio[edited_studio['Core/OAB'] == 'CORE']['Studio Hours'].sum()
                    st.metric("Core Hours", f"{core_hours:.2f}")
                with col3:
                    oab_hours = edited_studio[edited_studio['Core/OAB'] == 'OAB']['Studio Hours'].sum()
                    st.metric("OAB Hours", f"{oab_hours:.2f}")
                with col4:
                    projects_with_hours = (edited_studio['Studio Hours'] > 0).sum()
                    st.metric("Projects with Hours", projects_with_hours)
        else:
            st.info("Please process Production Files first to load project data")
    
    with tab4:
        st.header("🔍 Data Review")
        
        if not st.session_state.studio_data.empty:
            st.subheader("Studio Data Preview")
            st.dataframe(st.session_state.studio_data, use_container_width=True)
            
            # Download button for studio data
            csv = st.session_state.studio_data.to_csv(index=False)
            st.download_button(
                "📥 Download Studio Data (CSV)",
                csv,
                f"studio_data_{event_code}.csv",
                "text/csv",
                key='download-studio'
            )
        
        if not st.session_state.print_data.empty:
            st.subheader("Print Data Preview (first 100 rows)")
            st.dataframe(st.session_state.print_data.head(100), use_container_width=True)
            
            # Download button for print data
            csv = st.session_state.print_data.to_csv(index=False)
            st.download_button(
                "📥 Download Print Data (CSV)",
                csv,
                f"print_data_{event_code}.csv",
                "text/csv",
                key='download-print'
            )
    
    with tab5:
        st.header("💰 Cost Preview")
        
        if not st.session_state.studio_data.empty and not st.session_state.print_data.empty:
            # Calculate costs
            studio_df = st.session_state.studio_data.copy()
            print_df = st.session_state.print_data.copy()
            
            # Check if studio hours have been entered
            if studio_df['Studio Hours'].notna().any() and studio_df['Studio Hours'].sum() > 0:
                # Studio costs
                rate_map = {'Artwork': 49.5, 'Creative Artwork': 57, 'Digital': 49.5}
                studio_df['Rate'] = studio_df['Type'].map(rate_map).fillna(49.5)
                studio_df['Studio Cost'] = pd.to_numeric(studio_df['Studio Hours'], errors='coerce').fillna(0) * studio_df['Rate']
                
                # Print costs - using only Production Sell Price
                print_df['Production Sell Price'] = pd.to_numeric(print_df['Production Sell Price'], errors='coerce').fillna(0)
                print_df['Total including Spares'] = pd.to_numeric(print_df['Total including Spares'], errors='coerce').fillna(0)
                print_df['Total Cost'] = print_df['Production Sell Price'] * print_df['Total including Spares']
                
                # Assign Core/OAB to print items based on studio data
                project_core_oab = studio_df.set_index('Project Ref')['Core/OAB'].to_dict()
                print_df['Core/OAB'] = print_df['Project Ref'].map(project_core_oab).fillna('CORE')
                
                # Calculate totals
                studio_core = studio_df[studio_df['Core/OAB'] == 'CORE']['Studio Cost'].sum()
                studio_oab = studio_df[studio_df['Core/OAB'] == 'OAB']['Studio Cost'].sum()
                
                print_core = print_df[print_df['Core/OAB'] == 'CORE']['Total Cost'].sum()
                print_oab = print_df[print_df['Core/OAB'] == 'OAB']['Total Cost'].sum()
                
                # Display summary
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("### 💙 Core Costs")
                    st.metric("Studio", f"£{studio_core:,.2f}")
                    st.metric("Production", f"£{print_core:,.2f}")
                    st.metric("**Total Core**", f"**£{studio_core + print_core:,.2f}**")
                
                with col2:
                    st.markdown("### 💚 OAB Costs")
                    st.metric("Studio", f"£{studio_oab:,.2f}")
                    st.metric("Production", f"£{print_oab:,.2f}")
                    st.metric("**Total OAB**", f"**£{studio_oab + print_oab:,.2f}**")
                
                st.markdown("---")
                
                # Grand total
                grand_total = studio_core + studio_oab + print_core + print_oab
                st.metric("🎯 **Grand Total**", f"**£{grand_total:,.2f}**")
                
                # Breakdown by project
                with st.expander("📊 Breakdown by Project"):
                    project_summary = studio_df[['Project Ref', 'Project Description', 'Lines', 
                                                'Studio Hours', 'Studio Cost', 'Core/OAB']].copy()
                    
                    # Add production costs per project
                    print_by_project = print_df.groupby('Project Ref')['Total Cost'].sum().reset_index()
                    print_by_project.columns = ['Project Ref', 'Production Cost']
                    
                    project_summary = project_summary.merge(print_by_project, on='Project Ref', how='left')
                    project_summary['Production Cost'] = project_summary['Production Cost'].fillna(0)
                    project_summary['Total Cost'] = project_summary['Studio Cost'] + project_summary['Production Cost']
                    
                    st.dataframe(
                        project_summary.style.format({
                            'Studio Cost': '£{:,.2f}',
                            'Production Cost': '£{:,.2f}',
                            'Total Cost': '£{:,.2f}',
                            'Studio Hours': '{:.2f}'
                        }),
                        use_container_width=True
                    )
            else:
                st.warning("⚠️ Please enter Studio Hours in the Studio Hours Review tab to see cost preview")
        else:
            st.info("Process Production Files and add Studio Hours to see cost preview")
    
    with tab6:
        st.header("📄 Generate Invoice")
        
        ready_to_generate = (
            st.session_state.template_loaded and
            not st.session_state.studio_data.empty and
            not st.session_state.print_data.empty
        )
        
        if not ready_to_generate:
            st.warning("⚠️ Please complete the following steps before generating:")
            if not st.session_state.template_loaded:
                st.write("❌ Load template")
            if st.session_state.studio_data.empty:
                st.write("❌ Process production files")
        else:
            # Check if hours are present
            has_hours = st.session_state.studio_data['Studio Hours'].notna().any()
            
            if not has_hours:
                st.warning("⚠️ No Studio Hours found. The invoice will be generated but costs won't calculate without hours.")
            else:
                total_hours = st.session_state.studio_data['Studio Hours'].sum()
                st.success(f"✅ Ready to generate invoice with {total_hours:.2f} studio hours!")
            
            if st.button("🚀 Generate Invoice", type="primary", use_container_width=True):
                with st.spinner("Generating invoice with formatting..."):
                    try:
                        output_file, download_name, mime_type = generate_invoice(
                            st.session_state.template_info,
                            st.session_state.studio_data,
                            st.session_state.print_data,
                            event_name,
                            event_code
                        )

                        st.session_state.generated_file = output_file
                        st.success(f"✅ Invoice generated successfully!")

                        # Provide download button
                        with open(output_file, 'rb') as f:
                            excel_data = f.read()

                        st.download_button(
                            label="📥 Download Invoice",
                            data=excel_data,
                            file_name=download_name,
                            mime=mime_type,
                            use_container_width=True
                        )

                        st.info(f"Filename: {download_name}")
                        
                    except Exception as e:
                        st.error(f"Error generating invoice: {str(e)}")
                        st.exception(e)

else:
    # Welcome screen when no template is loaded
    st.info("👈 Please upload the Excel template in the sidebar to get started")
    
    st.markdown("""
    ## 🚀 Version 3.0 Features:
    
    - **Timesheet Integration**: Upload timesheet CSV to auto-populate studio hours
    - **Studio QC Exclusion**: Automatically excludes non-chargeable QC hours
    - **Smart Rounding**: Hours rounded up to nearest 0.25
    - **Production File Processing**: Handles multiple files with deduplication
    - **Event Code Generation**: Automatically converts "Event 10 2025" to "E1025"
    - **Format Preservation**: Maintains all Excel formatting and formulas
    - **Complete Workflow**: From raw data to finished invoice in minutes
    
    ## 📋 Workflow:
    
    1. **Upload Template** → Load your Excel invoice template
    2. **Process Production Files** → Upload Production Lines INTERNAL files
    3. **Import Timesheet** → Upload timesheet CSV to populate hours
    4. **Review Hours** → Edit studio hours and classifications as needed
    5. **Preview Costs** → Review cost breakdown before generating
    6. **Generate Invoice** → Create final Excel file with all formulas intact
    
    ## 🎯 Key Processing Rules:
    
    - **Studio QC hours excluded** (not chargeable)
    - **Hours rounded up** to nearest 0.25
    - **ROI projects** automatically set to OAB
    - **Duplicate Brief Refs** removed automatically
    - **Draft/unapproved items** filtered out
    """)

# Footer
st.markdown("---")
st.markdown("Built with ❤️ for Superdrug ITG Invoice Management | v3.0")
